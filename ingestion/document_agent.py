"""Document path: Excel / PDF / Word -> canonical tokens.

Explicit values (Excel cell, PDF table row) -> confidence 1.0 / not_required.
Prose- or swatch-inferred values go through the confidence gate (capped 0.8).
"""
import json
import logging
import os
import re

from ingestion import llm
from ingestion.schema_assembler import assemble, iter_tokens, make_token
from tools import pdf_table_extractor, swatch_detector

log = logging.getLogger(__name__)

_HEX = re.compile(r"#[0-9a-fA-F]{6}\b")
_DIM = re.compile(r"^\s*[0-9.]+(px|rem|em|pt)\s*$", re.I)

# Column/sheet name hints -> (category path, token type)
_CATEGORY_HINTS = {
    "color": (("color", "primitive"), "color"),
    "colour": (("color", "primitive"), "color"),
    "spacing": (("spacing",), "dimension"),
    "space": (("spacing",), "dimension"),
    "radius": (("radius",), "dimension"),
    "type": (("typography", "fontSize"), "fontSize"),
    "font": (("typography", "fontSize"), "fontSize"),
    "size": (("typography", "fontSize"), "fontSize"),
}


def _empty_groups():
    return {"color": {"primitive": {}}, "typography": {"fontSize": {}},
            "spacing": {}, "radius": {}}


def _place(groups, path, name, token):
    node = groups
    for key in path[:-1]:
        node = node.setdefault(key, {})
    node.setdefault(path[-1], {})[name] = token


def _category_for(header):
    for hint, dest in _CATEGORY_HINTS.items():
        if hint in (header or "").lower():
            return dest
    return None


# --- Excel: direct cell -> schema mapping -----------------------------------
def parse_excel(path):
    from openpyxl import load_workbook  # lazy
    wb = load_workbook(path, data_only=True)
    groups = _empty_groups()
    matched = False
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h or "").strip() for h in rows[0]]
        cats = [_category_for(h) or _category_for(ws.title) for h in headers]
        for row in rows[1:]:
            name = str(row[0]).strip() if row and row[0] is not None else None
            if not name:
                continue
            for col, cell in enumerate(row):
                if col == 0 or cell is None or col >= len(cats) or not cats[col]:
                    continue
                path_, ttype = cats[col]
                matched = True
                _place(groups, path_, name, make_token(
                    str(cell), ttype, confidence=1.0,
                    provenance={"source_page": f"{ws.title}!{name}"}))
    if not matched:
        raise ValueError("no recognizable token columns in workbook")
    return groups


# --- PDF: tables first, swatch fallback -------------------------------------
def parse_pdf(path, config):
    groups = _empty_groups()
    for table in pdf_table_extractor.extract_tables(path):
        for row in table:
            cells = [str(c).strip() for c in row if c]
            if len(cells) < 2:
                continue
            name, value = cells[0], cells[1]
            if _HEX.match(value):
                _place(groups, ("color", "primitive"), name,
                       make_token(_HEX.match(value).group(0).lower(), "color",
                                  provenance={"source_page": "pdf-table"}))
            elif _DIM.match(value):
                _place(groups, ("spacing",), name,
                       make_token(value, "dimension", provenance={"source_page": "pdf-table"}))
    if next(iter_tokens(groups), None):
        return groups
    return _swatch_fallback(path, config)  # no tables -> image swatches


def _swatch_fallback(path, config):
    groups = _empty_groups()
    # Render is out of scope here; expect a pre-rendered page image alongside the PDF.
    img = os.path.splitext(path)[0] + ".png"
    if not os.path.exists(img):
        log.warning("no tables and no %s for swatch fallback", img)
        return groups
    text = _safe_pdf_text(path)
    for i, hexv in enumerate(swatch_detector.dominant_colors(img)):
        name, conf = _name_swatch(hexv, text, config)
        _place(groups, ("color", "primitive"), f"{name or f'swatch-{i+1}'}",
               make_token(hexv, "color", confidence=conf, threshold=0.75,
                          provenance={"source_page": "pdf-swatch", "raw_observations": [hexv]}))
    return groups


def _name_swatch(hexv, text, config):
    if not llm.available():
        return None, 0.6
    try:
        role = llm.call(f"Color {hexv} appears near this spec text:\n{text[:1500]}\n"
                        f"Give one semantic token name (word only).",
                        api_key=(config or {}).get("google_key")).strip().split()[0]
        return re.sub(r"[^a-zA-Z0-9-]", "", role) or None, 0.75
    except llm.LLMUnavailable:
        return None, 0.6


# --- Word: prose -> tokens via single LLM call, confidence capped 0.8 -------
def parse_word(path, config):
    from docx import Document  # lazy
    text = "\n".join(p.text for p in Document(path).paragraphs if p.text.strip())
    return _prose_to_tokens(text, config)


def _prose_to_tokens(text, config):
    groups = _empty_groups()
    if not llm.available():
        log.warning("Word/prose path needs GOOGLE_API_KEY — no tokens extracted")
        return groups
    prompt = (
        "Extract design tokens from this spec. Return JSON: "
        '{"colors":{name:hex},"spacing":{name:value},"fontSizes":{name:value}}. '
        "Only include values explicitly stated.\n\n" + text[:6000])
    try:
        raw = llm.call(prompt, api_key=(config or {}).get("google_key"))
    except llm.LLMUnavailable:
        return groups
    data = _loads(raw)
    for name, hexv in (data.get("colors") or {}).items():
        _place(groups, ("color", "primitive"), name,
               make_token(hexv, "color", confidence=0.8, threshold=0.75,  # capped
                          description="prose-derived",
                          provenance={"source_page": "word-prose", "raw_observations": [str(hexv)]}))
    for name, val in (data.get("spacing") or {}).items():
        _place(groups, ("spacing",), name,
               make_token(str(val), "dimension", confidence=0.8, threshold=0.75,
                          description="prose-derived", provenance={"source_page": "word-prose"}))
    for name, val in (data.get("fontSizes") or {}).items():
        _place(groups, ("typography", "fontSize"), name,
               make_token(str(val), "fontSize", confidence=0.8, threshold=0.75,
                          description="prose-derived", provenance={"source_page": "word-prose"}))
    return groups


def _loads(raw):
    m = re.search(r"\{.*\}", raw or "", re.S)
    try:
        return json.loads(m.group(0)) if m else {}
    except json.JSONDecodeError:
        return {}


def _safe_pdf_text(path):
    try:
        return pdf_table_extractor.extract_text(path)
    except Exception:  # noqa: BLE001
        return ""


def run(file_path, config=None):
    config = config or {}
    ext = os.path.splitext(file_path)[1].lower()
    if ext in (".xlsx", ".xls"):
        groups = parse_excel(file_path)
    elif ext == ".pdf":
        groups = parse_pdf(file_path, config)
    elif ext in (".docx", ".doc"):
        groups = parse_word(file_path, config)
    else:
        raise ValueError(f"unsupported document type: {ext}")
    return assemble("document", os.path.basename(file_path), groups)
